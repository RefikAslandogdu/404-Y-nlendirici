from flask import Flask, render_template, request, jsonify, send_file
from urllib.parse import urlparse, unquote
from rapidfuzz import fuzz
import openpyxl
import tempfile
import os
import re

app = Flask(__name__)


def extract_keywords(url):
    """URL'den anlamlı anahtar kelimeleri çıkar."""
    parsed = urlparse(unquote(url))
    path = parsed.path.strip("/")

    # Uzantıları kaldır
    path = re.sub(r"\.(html?|php|aspx?|jsp)$", "", path, flags=re.IGNORECASE)

    # Ayırıcıları boşluğa çevir
    tokens = re.split(r"[-_/+.,]", path)

    # Boş ve çok kısa tokenları filtrele
    keywords = [t.lower() for t in tokens if len(t) > 1]
    return keywords


def extract_segments(url):
    """URL'den yol segmentlerini çıkar (kategori analizi için)."""
    parsed = urlparse(unquote(url))
    path = parsed.path.strip("/")
    path = re.sub(r"\.(html?|php|aspx?|jsp)$", "", path, flags=re.IGNORECASE)
    segments = [s.lower() for s in path.split("/") if s]
    return segments


def find_homepage(active_urls):
    """Aktif URL'ler arasından anasayfayı bul."""
    for url in active_urls:
        path = urlparse(url).path.strip("/")
        if not path or path in ("index", "index.html", "anasayfa", "home"):
            return url
    # Anasayfa bulunamazsa en kısa path'li URL'yi döndür
    return min(active_urls, key=lambda u: len(urlparse(u).path.strip("/")))


def score_match(source_url, target_url):
    """İki URL arasındaki benzerlik skorunu hesapla."""
    src_path = urlparse(unquote(source_url)).path.strip("/").lower()
    tgt_path = urlparse(unquote(target_url)).path.strip("/").lower()

    # 1) Tam yol eşleşmesi
    if src_path == tgt_path:
        return 100.0

    # 2) Kategori/segment eşleşmesi
    src_segments = extract_segments(source_url)
    tgt_segments = extract_segments(target_url)

    # Ortak segment sayısı (kategori mantığı)
    common_segments = sum(1 for s in src_segments if s in tgt_segments)
    max_segments = max(len(src_segments), len(tgt_segments), 1)
    category_score = (common_segments / max_segments) * 100

    # İlk segment (ana kategori) eşleşirse bonus
    category_bonus = 0
    if src_segments and tgt_segments and src_segments[0] == tgt_segments[0]:
        category_bonus = 25

    # 3) Fuzzy string benzerliği (yol bazlı)
    path_score = fuzz.token_sort_ratio(src_path, tgt_path)

    # 4) Anahtar kelime kesişimi
    src_kw = set(extract_keywords(source_url))
    tgt_kw = set(extract_keywords(target_url))

    if src_kw and tgt_kw:
        intersection = src_kw & tgt_kw
        union = src_kw | tgt_kw
        keyword_score = (len(intersection) / len(union)) * 100 if union else 0
    else:
        keyword_score = 0

    # Hiç ortak segment veya anahtar kelime yoksa skoru sınırla
    has_any_overlap = common_segments > 0 or len(src_kw & tgt_kw) > 0
    if not has_any_overlap:
        # Sadece fuzzy benzerlik var, mantıksal bağ yok — düşük tut
        return round(path_score * 0.2, 1)

    # Ağırlıklı ortalama: kategori mantığı ön planda
    final_score = (
        category_score * 0.3
        + category_bonus * 0.2
        + path_score * 0.25
        + keyword_score * 0.25
    )
    return round(min(final_score, 100.0), 1)


# Eşleşme kalitesinin yeterli sayılması için minimum skor
MATCH_THRESHOLD = 35


def find_best_match(redirect_url, active_urls):
    """404 URL için en iyi eşleşen aktif URL'yi bul.
    Yeterli eşleşme yoksa anasayfaya yönlendir."""
    best_url = None
    best_score = 0

    for active_url in active_urls:
        score = score_match(redirect_url, active_url)
        if score > best_score:
            best_score = score
            best_url = active_url

    # Skor eşiğin altındaysa anasayfaya yönlendir
    if best_score < MATCH_THRESHOLD:
        homepage = find_homepage(active_urls)
        return homepage, best_score, True

    return best_url, best_score, False


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    data = request.get_json()
    active_urls = [u.strip() for u in data.get("active_urls", []) if u.strip()]
    redirect_urls = [u.strip() for u in data.get("redirect_urls", []) if u.strip()]

    if not active_urls or not redirect_urls:
        return jsonify({"error": "Her iki listeye de URL girmelisiniz."}), 400

    results = []
    for rurl in redirect_urls:
        best_url, score, is_homepage = find_best_match(rurl, active_urls)
        results.append({
            "redirect_url": rurl,
            "target_url": best_url or "-",
            "score": score,
            "is_homepage": is_homepage,
        })

    return jsonify({"results": results})


@app.route("/export", methods=["POST"])
def export():
    data = request.get_json()
    results = data.get("results", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yönlendirmeler"

    # Başlıklar
    headers = ["404 URL (Yönlendirilecek)", "Hedef URL (Yönlenecek)", "Benzerlik Skoru"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Veriler
    for i, row in enumerate(results, 2):
        ws.cell(row=i, column=1, value=row["redirect_url"])
        ws.cell(row=i, column=2, value=row["target_url"])
        ws.cell(row=i, column=3, value=row["score"])

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="yonlendirmeler.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5859)
